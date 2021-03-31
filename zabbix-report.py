#!/usr/bin/env python3

# Core lib
import openpyxl
import sys
import json
import traceback
from os import environ
from time import sleep

from pyzabbix import ZabbixAPI


def json_sheets(json_data, name):
    data = json.loads(json_data)
    book = openpyxl.Workbook()
    del book['Sheet']
    for key in data:
        if isinstance(data[key], list):
            book.create_sheet(key)
            sheet = book[key]
            n = 1
            for r in data[key]:
                m = 1
                for c in r:
                    if n == 1:
                        for c1 in r:
                            sheet.cell(row=n, column=m).value = c1
                            m += 1
                        m = 1
                        n += 1
                    sheet.cell(row=n, column=m).value = r[c]
                    m += 1
                n += 1
    book.save(name)

def login_zabbix(zurl, zuser, zpass):
    zapi = ZabbixAPI(zurl)
    zapi.login(zuser, zpass)
    return zapi

def collect_zabbix(zapi):
    nvlAlerta = {'0': 'Not Classified', '1': 'Information',
                 '2': 'Warning', '3': 'Average', '4': 'High', '5': 'Disaster'}
    hostNames = []
    hostIds = []
    monitStatus = []
    agentStatus = []
    hostAux = []
    templates = []
    hostTrigger = []
    triggers = []
    severidade = []
    hostNameGroup = []
    hostGroup = []
    #Coletando Hosts e verificando se o Agent esta disponivel
    for host in zapi.host.get(output='extend'):
        if host['host'] != 'Zabbix server':
            hostNames.append(host['host'])
            hostIds.append(host['hostid'])
            agent = (host['available'])
            if "2" in agent:
                agentStatus.append('Agent com problema')
            elif "0" in agent:
                agentStatus.append('Agent nao startado')
            else:
                agentStatus.append('Monitorado')
            status = (host['status'])
            if "0" in status:
                monitStatus.append('Enable')
            else:
                monitStatus.append('Disable')
    #coletando Templates baseados no nome do Host
    for hostName in hostNames:
        for template in zapi.host.get(selectParentTemplates={'parentTemplates': 'name'}, filter={'host': hostName}):
            if len(template['parentTemplates']) == 0:
                hostAux.append(hostName)
                templates.append('Sem Templates')
            else:
                for index in range(0, len(template['parentTemplates'])):
                    hostAux.append(hostName)
                    templates.append(
                        template['parentTemplates'][index]['name'])
    #coletando as Triggers Descricoes e severidades baseadas no nome dos hosts
    for host in hostNames:
        for trigger in zapi.trigger.get(output='extend',expandDescription='extend', filter={'host': host}):
            hostTrigger.append(host)
            triggers.append(trigger['description'])
            severidade.append(nvlAlerta[trigger['priority']])
    #coletar hostgroups baseados em Hostnames
    for host in hostNames:
        for group in zapi.host.get(selectGroups='extend', filter={'host': host}):
            for index in range(0, len(group['groups'])):
                hostNameGroup.append(host)
                hostGroup.append(group['groups'][index]['name'])
    #formatando output
    output = {
        'templates': [],
        'status': [],
        'triggers': [],
        'hostgroups': []
    }
    for i, host in enumerate(hostAux):
        output['templates'].append(
            {'Hostname': host, 'Template': templates[i]})

    for i, host in enumerate(hostNames):
        output['status'].append(
            {'Hostname': host, 'Status': monitStatus[i], 'Agent': agentStatus[i]})
    for i, host in enumerate(hostTrigger):
        output['triggers'].append(
            {'Hostname': host, 'Triggers': triggers[i], 'Severity': severidade[i]})
    for i, host in enumerate(hostNameGroup):
        output['hostgroups'].append(
            {'Hostname': host, 'Hostgroup': hostGroup[i]})

    return output


def verify_parameters(event, params):
    body = dict()
    headers = dict()
    query = dict()
    if event.get('body') and event['body']:
        body = json.loads(event['body'])

    if event.get('headers') and event['headers']:
        headers = event['headers']

    if event.get('queryStringParameters') and event['queryStringParameters']:
        query = event['queryStringParameters']

    for k, p in params.items():
        if headers.get(k):
            params[k] = headers[k]
        elif body.get(k):
            params[k] = body[k]
        elif query.get(k):
            params[k] = query[k]
        elif event.get(k):
            params[k] = event[k]

    return params


def final_response(result):
    response = {
        'statusCode': 200,
        'body': json.dumps(result, sort_keys=True, indent=4)
    }
    return response


def lambda_handler(event, context):
    params = verify_parameters(
        event, {'zurl': None, 'zuser': None, 'zpass': None})
    zapi = login_zabbix(params['zurl'], params['zuser'], params['zpass'])
    return final_response(collect_zabbix(zapi))

if __name__ == '__main__':
    json_sheets(lambda_handler({
        'zurl': sys.argv[1], 'zuser': sys.argv[2], 'zpass': sys.argv[3]
    }, None).get('body'), '{}.xlsx'.format(sys.argv[4]))
